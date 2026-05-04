import csv
import io
from django.contrib import admin
from projects.models.Projects import Projects
from ..models.UserLedger import UserLedger
from ..models.UserLedgerTransaction import UserLedgerTransaction
from common.filters.adminModelFilter import TableForiegnKeyListFilter
from django.db.models import Q
from django.db.models import F
from num2words import num2words
from django.urls import path, reverse
from ..utils import format_indian_currency, amount_to_words
from django.utils.html import format_html
from django.http import HttpResponseRedirect, HttpResponse
from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
try:
    import weasyprint
except Exception:
    weasyprint = None
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

class UserLedgerTransactionsInline(admin.TabularInline):
    model = UserLedgerTransaction
    extra = 1
    exclude = ('created_at', 'updated_at')
    show_change_link = True
    verbose_name = "User Ledger Transaction"
    verbose_name_plural = "User Ledger Transactions"
    template = 'admin/accounts/user_ledger/tabular_paginated.html'

    readonly_fields = [
        'amount_display', 'amount_words',
        'paid_amount_display', 'paid_amount_words',
        'balance_display', 'balance_words'
    ]

    fields = ['paid_on', 'payment_type', 'amount', 'amount_display', 'amount_words', 'detail']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Last 5 transactions by date, displayed in ascending order
        last_5_ids = list(
            qs.order_by(F('paid_on').desc(nulls_last=True)).values_list('id', flat=True)[:5]
        )
        return qs.filter(id__in=last_5_ids).order_by(F('paid_on').asc(nulls_last=True))

    def amount_display(self, obj):
        if obj.id:
            return format_indian_currency(obj.amount)
        return "-"
    amount_display.short_description = "Amount (Fmt)"

    def amount_words(self, obj):
        if obj.id:
            return amount_to_words(obj.amount)
        return "-"
    amount_words.short_description = "Words"

    def paid_amount_display(self, obj):
        if obj.id and obj.user_ledger:
            paid_amt = obj.user_ledger.amount - obj.user_ledger.balance
            return format_indian_currency(paid_amt)
        return "-"
    paid_amount_display.short_description = "Paid (Fmt)"

    def paid_amount_words(self, obj):
        if obj.id and obj.user_ledger:
            paid_amt = obj.user_ledger.amount - obj.user_ledger.balance
            return amount_to_words(paid_amt)
        return "-"
    paid_amount_words.short_description = "Paid Words"

    def balance_display(self, obj):
        if obj.id and obj.user_ledger:
            return format_indian_currency(obj.user_ledger.balance)
        return "-"
    balance_display.short_description = "Bal (Fmt)"

    def balance_words(self, obj):
        if obj.id and obj.user_ledger:
            return amount_to_words(obj.user_ledger.balance)
        return "-"
    balance_words.short_description = "Bal Words"

    def has_add_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


class UserLedgerTransactionViewInline(admin.TabularInline):
    model = UserLedgerTransaction
    extra = 0
    can_delete = False
    show_change_link = True
    verbose_name = "Transaction"
    verbose_name_plural = "Transactions View"
    template = 'admin/accounts/user_ledger/tabular_paginated_view.html'

    readonly_fields = ['paid_on', 'payment_type', 'amount', 'amount_display', 'amount_words', 'detail']
    fields = ['paid_on', 'payment_type', 'amount', 'amount_display', 'amount_words', 'detail']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by(F('paid_on').asc(nulls_last=True))

    def get_formset(self, request, obj=None, **kwargs):
        FormSet = super().get_formset(request, obj, **kwargs)

        class PaginatedInlineFormSet(FormSet):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                qs = self.queryset
                paginator = Paginator(qs, 15)
                page_num = request.GET.get('trans_page', 1)
                try:
                    self.page_obj = paginator.page(page_num)
                except (PageNotAnInteger, EmptyPage):
                    self.page_obj = paginator.page(1)
                self.queryset = self.page_obj.object_list

        return PaginatedInlineFormSet

    def amount_display(self, obj):
        if obj.id:
            return format_indian_currency(obj.amount)
        return "-"
    amount_display.short_description = "Amount (Fmt)"

    def amount_words(self, obj):
        if obj.id:
            return amount_to_words(obj.amount)
        return "-"
    amount_words.short_description = "Words"

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

class UserLedgerAdmin(admin.ModelAdmin):
    list_select_related = True
    list_display = ['creditor_name', 'debtor_name', 'project_id', 'formatted_amount', 'formatted_balance', 'paid_amount', 'ledger_actions' ] 
    exclude = ('created_at', 'updated_at')
    readonly_fields = ['paid_amount']

    list_per_page = 15          # ← yeh add karo
    list_max_show_all = 100     # ← yeh add karo

    def creditor_name(self, obj):
        if obj.creditor.first_name: 
            return obj.creditor.first_name + " " + obj.creditor.last_name
        return obj.creditor
    
    def debtor_name(self, obj):
        if obj.debtor.first_name: 
            return obj.debtor.first_name + " " + obj.debtor.last_name
        return obj.debtor

    def formatted_amount(self, obj):
        if obj.amount:
            words = amount_to_words(obj.amount)
            return format_html("{}<br><small>({})</small>", format_indian_currency(obj.amount), words)
        return "-"
    formatted_amount.short_description = "Amount"

    def formatted_balance(self, obj):
        if obj.balance:
            words = amount_to_words(obj.balance)
            return format_html("{}<br><small>({})</small>", format_indian_currency(obj.balance), words)
        return "-"
    formatted_balance.short_description = "Balance"

    def paid_amount(self, obj):
        if obj.amount and obj.balance:
            paidAmount = obj.amount - obj.balance
            words = amount_to_words(paidAmount)
            return format_html("{}<br><small>({})</small>", format_indian_currency(paidAmount), words)
        return "-"
    paid_amount.short_description = "Paid Amount"

    inlines = [UserLedgerTransactionsInline, UserLedgerTransactionViewInline]
    search_fields = ['amount', 'balance']
    list_filter = [TableForiegnKeyListFilter("Projects", "project_id", "name", Projects)]

    # ─── Actions column with Download dropdown ─────────────────────────────

    def ledger_actions(self, obj):
        pdf_url   = reverse('admin:download_user_ledger',       args=[obj.pk])
        excel_url = reverse('admin:download_user_ledger_excel', args=[obj.pk])
        csv_url   = reverse('admin:download_user_ledger_csv',   args=[obj.pk])
        uid = f"ldrop_{obj.pk}"
        return format_html(
            '<div style="display:inline-block;">'
            '<button type="button" data-sved-toggle="1"'
            ' onclick="svedToggleDrop(this,\'{uid}\')"'
            ' style="background:#0d6efd;color:#fff;padding:5px 12px;border:none;'
            'border-radius:4px;font-size:12px;cursor:pointer;white-space:nowrap;">'
            '&#128196; Download &#9660;</button>'
            '<div id="{uid}" class="sved-ddmenu"'
            ' style="display:none;position:fixed;background:#fff;'
            'border:1px solid #dee2e6;border-radius:4px;padding:4px 0;'
            'min-width:130px;z-index:99999;box-shadow:0 4px 12px rgba(0,0,0,.15);">'
            '<a href="{pdf}" target="_blank"'
            ' style="display:block;padding:6px 14px;color:#212529;text-decoration:none;font-size:12px;">&#128196; PDF</a>'
            '<a href="{excel}" target="_blank"'
            ' style="display:block;padding:6px 14px;color:#212529;text-decoration:none;font-size:12px;">&#128202; Excel</a>'
            '<a href="{csv}" target="_blank"'
            ' style="display:block;padding:6px 14px;color:#212529;text-decoration:none;font-size:12px;">&#128203; CSV</a>'
            '</div></div>',
            uid=uid, pdf=pdf_url, excel=excel_url, csv=csv_url,
        )
    ledger_actions.short_description = "Actions"
    ledger_actions.allow_tags = True
    

    change_form_template = 'admin/accounts/user_ledger/change_form.html'

    class Media:
        js = ('admin/js/amountFormat1.js', 'admin/js/ledger_dropdown.js')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        user = request.user
        if user.is_superuser:
            return qs
        return qs.filter(Q(creditor=user) | Q(debtor=user)).distinct()
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'user-ledger/<int:ledger_id>/pdf/',
                self.admin_site.admin_view(self.download_ledger_pdf),
                name='download_user_ledger',
            ),
            path(
                'user-ledger/<int:ledger_id>/excel/',
                self.admin_site.admin_view(self.download_ledger_excel),
                name='download_user_ledger_excel',
            ),
            path(
                'user-ledger/<int:ledger_id>/csv/',
                self.admin_site.admin_view(self.download_ledger_csv),
                name='download_user_ledger_csv',
            ),
        ]
        return custom_urls + urls
    
    def _get_ledger_context(self, ledger_id):
        ledger = UserLedger.objects.select_related().get(pk=ledger_id)
        transactions = list(
            UserLedgerTransaction.objects.filter(user_ledger=ledger)
            .order_by(F('paid_on').asc(nulls_last=True))
        )
        for txn in transactions:
            txn.formatted_amount = format_indian_currency(txn.amount)
        paid_amount = (ledger.amount - ledger.balance) if (ledger.amount and ledger.balance) else 0
        paid_words  = num2words(int(paid_amount), lang='en_IN').title() if paid_amount else ""
        return ledger, transactions, paid_amount, paid_words

    def download_ledger_pdf(self, request, ledger_id):
        ledger, transactions, paid_amount, paid_words = self._get_ledger_context(ledger_id)
        context = {
            'ledger':            ledger,
            'transactions':      transactions,
            'paid_amount':       paid_amount,
            'paid_words':        paid_words,
            'formatted_amount':  format_indian_currency(ledger.amount),
            'formatted_balance': format_indian_currency(ledger.balance),
            'formatted_paid':    format_indian_currency(paid_amount),
        }
        if weasyprint is None:
            return HttpResponse("PDF generation is unavailable (weasyprint not installed).", status=503)
        html_string = render_to_string('admin/accounts/user_ledger_pdf.html', context)
        pdf_file    = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="ledger_{ledger_id}.pdf"'
        return response

    def download_ledger_excel(self, request, ledger_id):
        if openpyxl is None:
            return HttpResponse("Excel generation is unavailable (openpyxl not installed).", status=503)

        ledger, transactions, paid_amount, _ = self._get_ledger_context(ledger_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Ledger"

        # ── styles ──
        header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill   = PatternFill('solid', fgColor='1A56DB')
        center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        thin_border   = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        title_font    = Font(name='Calibri', bold=True, size=14, color='1A56DB')
        label_font    = Font(name='Calibri', bold=True, size=10)
        value_font    = Font(name='Calibri', size=10)
        blue_fill     = PatternFill('solid', fgColor='EFF6FF')

        def _name(user):
            return f"{user.first_name} {user.last_name}".strip() or str(user)

        creditor_name = _name(ledger.creditor)
        debtor_name   = _name(ledger.debtor)
        project_name  = str(ledger.project_id) if ledger.project_id else "—"

        # ── Row 1: Title ──
        ws.merge_cells('A1:F1')
        ws['A1'] = "SVED — User Ledger Statement"
        ws['A1'].font      = title_font
        ws['A1'].alignment = center_align

        # ── Rows 3-5: Ledger Info ──
        info = [
            ("Creditor",      creditor_name),
            ("Debtor",        debtor_name),
            ("Project",       project_name),
            ("Total Amount",  format_indian_currency(ledger.amount)),
            ("Paid Amount",   format_indian_currency(paid_amount)),
            ("Balance",       format_indian_currency(ledger.balance)),
        ]
        for i, (lbl, val) in enumerate(info, start=3):
            ws.cell(row=i, column=1, value=lbl).font      = label_font
            ws.cell(row=i, column=1).fill                 = blue_fill
            ws.cell(row=i, column=2, value=val).font      = value_font
            ws.cell(row=i, column=2).alignment            = left_align

        # ── Row 10: Transaction table header ──
        txn_header_row = 10
        txn_cols = ['#', 'Paid On', 'Payment Type', 'Detail', 'Amount (₹)']
        for col_idx, col_name in enumerate(txn_cols, start=1):
            cell = ws.cell(row=txn_header_row, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        # ── Transaction rows ──
        for row_idx, txn in enumerate(transactions, start=1):
            row_num = txn_header_row + row_idx
            row_fill = PatternFill('solid', fgColor='F5F8FF') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            row_data = [
                row_idx,
                txn.paid_on.strftime('%d-%m-%Y') if txn.paid_on else '—',
                txn.payment_type or '—',
                txn.detail or '—',
                float(txn.amount) if txn.amount else 0,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.border    = thin_border
                cell.alignment = center_align if col_idx in (1, 2, 3, 5) else left_align

        # ── Column widths ──
        col_widths = [6, 14, 16, 40, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="ledger_{ledger_id}.xlsx"'
        return response

    def download_ledger_csv(self, request, ledger_id):
        ledger, transactions, paid_amount, _ = self._get_ledger_context(ledger_id)

        def _name(user):
            return f"{user.first_name} {user.last_name}".strip() or str(user)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="ledger_{ledger_id}.csv"'
        response.write('﻿')  # UTF-8 BOM so Excel opens Hindi/special chars correctly

        writer = csv.writer(response)

        # Ledger summary block
        writer.writerow(['SVED - User Ledger Statement'])
        writer.writerow([])
        writer.writerow(['Creditor', _name(ledger.creditor)])
        writer.writerow(['Debtor',   _name(ledger.debtor)])
        writer.writerow(['Project',  str(ledger.project_id) if ledger.project_id else '—'])
        writer.writerow(['Total Amount', format_indian_currency(ledger.amount)])
        writer.writerow(['Paid Amount',  format_indian_currency(paid_amount)])
        writer.writerow(['Balance',      format_indian_currency(ledger.balance)])
        writer.writerow([])

        # Transactions
        writer.writerow(['#', 'Paid On', 'Payment Type', 'Detail', 'Amount'])
        for i, txn in enumerate(transactions, start=1):
            writer.writerow([
                i,
                txn.paid_on.strftime('%d-%m-%Y') if txn.paid_on else '—',
                txn.payment_type or '—',
                txn.detail or '—',
                float(txn.amount) if txn.amount else 0,
            ])

        return response

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        if object_id:
            extra_context['download_pdf_url']   = reverse('admin:download_user_ledger',       args=[object_id])
            extra_context['download_excel_url']  = reverse('admin:download_user_ledger_excel', args=[object_id])
            extra_context['download_csv_url']    = reverse('admin:download_user_ledger_csv',   args=[object_id])
        return super().changeform_view(request, object_id, form_url, extra_context)


admin.site.register(UserLedger, UserLedgerAdmin)
