import csv
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from django.http import HttpResponse
from django.utils import timezone
from django.template.loader import render_to_string
from .models.Customers import Customers
from .models.CustomerLedger import CustomerLedger
from .models.CustomerLedgerTransaction import CustomerLedgerTransaction
from common.utils.format_currency import format_indian_currency
from num2words import num2words
from django.forms import modelform_factory
from customers.models.CustomerRequestTransaction import CustomerRequestTransaction

try:
    import weasyprint
except Exception:
    weasyprint = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


def customer_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            customer = Customers.objects.get(username=username)
        except Customers.DoesNotExist:
            customer = None

        if customer:
            if customer.status != Customers.STATUS_ACTIVE:
                messages.error(request, "Account is inactive. Contact to superadmin")
                return redirect('login')

            if check_password(password, customer.password_hash):
                request.session['customer_id'] = customer.id
                request.session['customer_name'] = customer.username
                customer.last_login_on = timezone.now()
                customer.save(update_fields=['last_login_on'])
                return redirect('dashboard')
            else:
                messages.error(request, "Wrong password")
        else:
            messages.error(request, "Invalid username and password")

    return render(request, 'frontend/customer/login.html')

def customer_logout(request):
    request.session.flush()
    return redirect('login')

def customer_dashboard(request):
    if not request.session.get('customer_id'):
        return redirect('login')

    customer_id = request.session.get('customer_id')
    customer = Customers.objects.get(id=customer_id)
    Customers.objects.filter(pk=customer.pk).update(last_active_on=timezone.now())

    customerLedger = CustomerLedger.objects.filter(customer_id=customer).first()

    if not customerLedger:
        messages.error(request, "No ledger found for your account. Please contact admin.")
        return render(request, 'frontend/customer/dashboard.html', {
            'customer': customer,
            'name': (customer.first_name or '') + ' ' + (customer.last_name or ''),
        })

    customerTransactions = customerLedger.customer_ledger_transactions.all().order_by('-paid_on')

    total_amount_words = num2words(customerLedger.amount, lang='en_IN').title()
    balance_words = num2words(customerLedger.balance, lang='en_IN').title()
    paid_amount = customerLedger.amount - customerLedger.balance
    paid_amount_words = num2words(paid_amount, lang='en_IN').title()

    context = {
        'customer': customer,
        'name': (customer.first_name or '') + ' ' + (customer.last_name or ''),
        'email': customer.email,
        'total_amount': format_indian_currency(customerLedger.amount),
        'total_amount_words': total_amount_words,
        'balance': format_indian_currency(customerLedger.balance),
        'balance_words': balance_words,
        'paid_amount': format_indian_currency(paid_amount),
        'paid_amount_words': paid_amount_words,
        'transactions': customerTransactions,
        'house_no': customerLedger.project_house_id.plot_no,
        'format_indian_currency': format_indian_currency,
    }

    return render(request, 'frontend/customer/dashboard.html', context)

def customer_request_list(request):
    if not request.session.get('customer_id'):
        return redirect('login')
    
    customer_id = request.session.get('customer_id')
    customer = Customers.objects.get(id=customer_id)

    customerLedger = CustomerLedger.objects.filter(customer_id=customer).first()

    if not customerLedger:
        messages.error(request, "No ledger found for your account. Please contact admin.")
        return render(request, 'frontend/customer/request_listing.html', {
            'customer': customer,
            'name': (customer.first_name or '') + ' ' + (customer.last_name or ''),
            'transactions': [],
        })

    customerRequestTransactions = customerLedger.customer_transactions.filter(
        status=CustomerRequestTransaction.STATUS_NEW
    ).order_by('-paid_on')

    total_amount_words = num2words(customerLedger.amount, lang='en_IN').title()
    balance_words = num2words(customerLedger.balance, lang='en_IN').title()
    paid_amount = customerLedger.amount - customerLedger.balance
    paid_amount_words = num2words(paid_amount, lang='en_IN').title()

    context = {
        'customer': customer,
        'name': (customer.first_name or '') + ' ' + (customer.last_name or ''),
        'email': customer.email,
        'total_amount': format_indian_currency(customerLedger.amount),
        'total_amount_words': total_amount_words,
        'balance': format_indian_currency(customerLedger.balance),
        'balance_words': balance_words,
        'paid_amount': format_indian_currency(paid_amount),
        'paid_amount_words': paid_amount_words,
        'transactions': customerRequestTransactions,
        'house_no': customerLedger.project_house_id.plot_no,
        'format_indian_currency': format_indian_currency,
    }

    return render(request, 'frontend/customer/request_listing.html', context)


def customer_request_dashboard(request):

    if not request.session.get('customer_id'):
        return redirect('login')
    
    customer_id = request.session.get('customer_id')
    customer = Customers.objects.get(id=customer_id)

    customerLedger = CustomerLedger.objects.filter(customer_id=customer).first()

    CustomerRequestTransactionForm = modelform_factory(
        CustomerRequestTransaction,
        exclude=['created_at', 'updated_at','customer_ledger','status']
    )

    if request.method == 'POST':
        form = CustomerRequestTransactionForm(request.POST)
    else:
        form = CustomerRequestTransactionForm()

    for field_name, field in form.fields.items():

        if field.widget.__class__.__name__ in ['CheckboxInput']:
            field.widget.attrs['class'] = 'form-check-input'
        else:
            field.widget.attrs['class'] = 'form-control'

        if field_name == 'paid_on':
            field.widget.input_type = 'date'
            field.widget.attrs['class'] = 'form-control'

    if request.method == 'POST':
        if form.is_valid():
            obj = form.save(commit=False)
            obj.customer_ledger = customerLedger

            obj.save()
            messages.success(request, "Your request is submitted successfully.")
            return redirect('/customer/requests/')

    context = {
        'form': form
    }

    return render(request, 'frontend/customer/request.html', context)


# ── Customer portal: download ledger ────────────────────────────────────────

def _get_download_data(request):
    customer_id = request.session.get('customer_id')
    if not customer_id:
        return None
    customer = Customers.objects.get(id=customer_id)
    ledger = CustomerLedger.objects.select_related(
        'customer_id', 'project_id', 'project_house_id'
    ).filter(customer_id=customer).first()
    if not ledger:
        return None
    transactions = list(
        CustomerLedgerTransaction.objects.filter(customer_ledger=ledger)
        .select_related('paid_to')
        .order_by('paid_on')
    )
    for txn in transactions:
        txn.formatted_amount = format_indian_currency(txn.amount)
    paid_amount = (ledger.amount - ledger.balance) if (ledger.amount and ledger.balance) else 0
    paid_words  = num2words(int(paid_amount), lang='en_IN').title() if paid_amount else ""
    return customer, ledger, transactions, paid_amount, paid_words


def customer_download_pdf(request):
    data = _get_download_data(request)
    if not data:
        return redirect('login')
    customer, ledger, transactions, paid_amount, paid_words = data
    if weasyprint is None:
        return HttpResponse("PDF generation is unavailable.", status=503)
    context = {
        'ledger':            ledger,
        'transactions':      transactions,
        'paid_amount':       paid_amount,
        'paid_words':        paid_words,
        'formatted_amount':  format_indian_currency(ledger.amount),
        'formatted_balance': format_indian_currency(ledger.balance),
        'formatted_paid':    format_indian_currency(paid_amount),
    }
    html_string = render_to_string('admin/customers/customer_ledger_pdf.html', context)
    pdf_file = weasyprint.HTML(
        string=html_string, base_url=request.build_absolute_uri('/')
    ).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="my_ledger.pdf"'
    return response


def customer_download_excel(request):
    data = _get_download_data(request)
    if not data:
        return redirect('login')
    if openpyxl is None:
        return HttpResponse("Excel generation is unavailable.", status=503)
    customer, ledger, transactions, paid_amount, _ = data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Ledger"

    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', fgColor='1A56DB')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    title_font = Font(name='Calibri', bold=True, size=14, color='1A56DB')
    label_font = Font(name='Calibri', bold=True, size=10)
    value_font = Font(name='Calibri', size=10)
    blue_fill  = PatternFill('solid', fgColor='EFF6FF')

    customer_name = f"{customer.first_name} {customer.last_name}".strip() or customer.username

    ws.merge_cells('A1:F1')
    ws['A1'] = "SVED — My Ledger Statement"
    ws['A1'].font      = title_font
    ws['A1'].alignment = center_align

    info = [
        ("Customer",     customer_name),
        ("Project",      str(ledger.project_id) if ledger.project_id else "—"),
        ("House",        str(ledger.project_house_id) if ledger.project_house_id else "—"),
        ("Total Amount", format_indian_currency(ledger.amount)),
        ("Paid Amount",  format_indian_currency(paid_amount)),
        ("Balance",      format_indian_currency(ledger.balance)),
    ]
    for i, (lbl, val) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=lbl).font = label_font
        ws.cell(row=i, column=1).fill            = blue_fill
        ws.cell(row=i, column=2, value=val).font = value_font
        ws.cell(row=i, column=2).alignment       = left_align

    hdr_row = 10
    for col_idx, col_name in enumerate(
        ['#', 'Paid On', 'Payment Type', 'Paid To', 'Detail', 'Amount (₹)'], start=1
    ):
        cell = ws.cell(row=hdr_row, column=col_idx, value=col_name)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border

    for row_idx, txn in enumerate(transactions, start=1):
        row_num  = hdr_row + row_idx
        row_fill = PatternFill('solid', fgColor='F5F8FF') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        paid_to  = ""
        if txn.paid_to:
            paid_to = f"{txn.paid_to.first_name} {txn.paid_to.last_name}".strip() or str(txn.paid_to)
        row_data = [
            row_idx,
            txn.paid_on.strftime('%d-%m-%Y') if txn.paid_on else '—',
            txn.payment_type or '—',
            paid_to or '—',
            txn.detail or '—',
            float(txn.amount) if txn.amount else 0,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = row_fill; cell.border = thin_border
            cell.alignment = center_align if col_idx in (1, 2, 3, 6) else left_align

    for i, w in enumerate([6, 14, 16, 20, 36, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="my_ledger.xlsx"'
    return response


def customer_download_csv(request):
    data = _get_download_data(request)
    if not data:
        return redirect('login')
    customer, ledger, transactions, paid_amount, _ = data

    customer_name = f"{customer.first_name} {customer.last_name}".strip() or customer.username

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="my_ledger.csv"'
    response.write('﻿')

    writer = csv.writer(response)
    writer.writerow(['SVED - My Ledger Statement'])
    writer.writerow([])
    writer.writerow(['Customer',     customer_name])
    writer.writerow(['Project',      str(ledger.project_id) if ledger.project_id else '—'])
    writer.writerow(['House',        str(ledger.project_house_id) if ledger.project_house_id else '—'])
    writer.writerow(['Total Amount', format_indian_currency(ledger.amount)])
    writer.writerow(['Paid Amount',  format_indian_currency(paid_amount)])
    writer.writerow(['Balance',      format_indian_currency(ledger.balance)])
    writer.writerow([])
    writer.writerow(['#', 'Paid On', 'Payment Type', 'Paid To', 'Detail', 'Amount'])
    for i, txn in enumerate(transactions, start=1):
        paid_to = ""
        if txn.paid_to:
            paid_to = f"{txn.paid_to.first_name} {txn.paid_to.last_name}".strip() or str(txn.paid_to)
        writer.writerow([
            i,
            txn.paid_on.strftime('%d-%m-%Y') if txn.paid_on else '—',
            txn.payment_type or '—',
            paid_to or '—',
            txn.detail or '—',
            float(txn.amount) if txn.amount else 0,
        ])
    return response