import csv
import io
from django.contrib import admin
from projects.models.Projects import Projects
from customers.models.CustomerLedger import CustomerLedger
from customers.models.CustomerLedgerTransaction import CustomerLedgerTransaction
from customers.models.CustomerRequestTransaction import CustomerRequestTransaction
from common.filters.adminModelFilter import TableForiegnKeyListFilter
from django.db.models import Q, F, Sum
from num2words import num2words
from common.utils.format_currency import format_indian_currency
from django.utils.html import format_html
from django.urls import path, reverse
from django.http import HttpResponseRedirect, HttpResponse
from django.db import transaction
from django.template.loader import render_to_string
try:
    import weasyprint
except Exception:
    weasyprint = None
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


class CustomerLedgerTransactionsInline(admin.TabularInline):
    model = CustomerLedgerTransaction
    extra = 2
    exclude = ('created_at', 'updated_at') 
    verbose_name = "Customer Ledger Transaction"
    verbose_name_plural = "Customer Ledger Transactions"
    fields = ['paid_on', 'payment_type', 'mode', 'amount', 'paid_to', 'detail']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by(F('paid_on').asc(nulls_last=True))

    def has_add_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class CustomerLedgerRequestInline(admin.TabularInline):
    model = CustomerRequestTransaction
    extra = 0
    exclude = ('created_at', 'updated_at') 
    verbose_name = "Customer Transaction Request"
    verbose_name_plural = "Customer Transaction Requests" 
    fields = [
        'paid_on',
        'payment_type',
        'mode',
        'amount',
        'paid_to',
        'detail',
        'action_button'
    ]

    readonly_fields = [
        'paid_on',
        'payment_type',
        'mode',
        'amount',
        'paid_to',
        'detail',
        'action_button'
    ]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.filter(status=10).order_by(F('paid_on').asc(nulls_last=True))
    
    def action_button(self, obj):
        url  = reverse('admin:accept_customer_request', args=[obj.id])
        url1 = reverse('admin:reject_customer_request', args=[obj.id])

        return format_html(
            '<a class="button" href="{}" '
            'style="background:#198754;color:white;padding:5px 10px;border-radius:4px;text-decoration:none;">'
            'Accept</a> '
            ' <a class="button" href="{}" '
            'style="background:red;color:white;padding:5px 10px;border-radius:4px;text-decoration:none;">'
            'Reject</a>',
            url, url1
        )

    action_button.short_description = 'Action'

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class CustomerLedgerAdmin(admin.ModelAdmin):

    list_select_related = True

    # ✅ 'ledger_actions' column add ki gayi
    list_display = [
        'customer_id',
        'project_id',
        'project_house_id',
        'formatted_amount',
        'formatted_loan_amount',
        'formatted_balance',
        'paid_amount',
        'ledger_actions',
    ]

    exclude = ('created_at', 'updated_at')
    readonly_fields = ['paid_amount', 'balance']

    list_per_page = 15          # ← yeh add karo
    list_max_show_all = 100     # ← yeh add karo

    # ─── existing display methods ───────────────────────────────────────────

    def formatted_amount(self, obj):
        return format_indian_currency(obj.amount)
    formatted_amount.short_description = "Amount"

    def formatted_loan_amount(self, obj):
        if obj.loan_amount:
            return format_indian_currency(obj.loan_amount)
        return "—"
    formatted_loan_amount.short_description = "Loan Amount"

    def formatted_balance(self, obj):
        return format_indian_currency(obj.balance)
    formatted_balance.short_description = "Balance"

    def paid_amount(self, obj):
        if obj.amount and obj.balance:
            paidAmount = obj.amount - obj.balance
            words = num2words(paidAmount, lang='en_IN').title()
            return format_html(
                "{}<br><small>({})</small>",
                format_indian_currency(paidAmount),
                words
            )
        return "-"
    paid_amount.short_description = "Paid Amount"

    # ─── Actions column with Download dropdown ─────────────────────────────

    def ledger_actions(self, obj):
        pdf_url   = reverse('admin:download_customer_ledger',       args=[obj.pk])
        excel_url = reverse('admin:download_customer_ledger_excel', args=[obj.pk])
        csv_url   = reverse('admin:download_customer_ledger_csv',   args=[obj.pk])
        recalc_url = reverse('admin:recalculate_customer_balance',  args=[obj.pk])
        uid = f"cldrop_{obj.pk}"
        return format_html(
            '<div style="display:inline-block;vertical-align:middle;">'
            '<button type="button" data-sved-toggle="1"'
            ' onclick="svedToggleDrop(this,\'{uid}\')"'
            ' style="background:#0d6efd;color:#fff;padding:5px 12px;border:none;'
            'border-radius:4px;font-size:12px;cursor:pointer;white-space:nowrap;">'
            '&#128196; Download &#9660;</button>'
            '<div id="{uid}" class="sved-ddmenu"'
            ' style="display:none;position:fixed;background:#fff;'
            'border:1px solid #dee2e6;border-radius:4px;padding:4px 0;'
            'min-width:130px;z-index:99999;box-shadow:0 4px 12px rgba(0,0,0,.15);">'
            '<a href="{pdf}" target="_blank"'
            ' style="display:block;padding:6px 14px;color:#212529;text-decoration:none;font-size:12px;">&#128196; PDF</a>'
            '<a href="{excel}" target="_blank"'
            ' style="display:block;padding:6px 14px;color:#212529;text-decoration:none;font-size:12px;">&#128202; Excel</a>'
            '<a href="{csv}" target="_blank"'
            ' style="display:block;padding:6px 14px;color:#212529;text-decoration:none;font-size:12px;">&#128203; CSV</a>'
            '</div></div>'
            '&nbsp;<button type="button"'
            ' onclick="if(confirm(\'Is customer ka balance transactions se recalculate karke update karna chahte ho?\'))'
            '{{window.location=\'{recalc}\'}}"'
            ' style="background:#198754;color:#fff;padding:5px 10px;border:none;'
            'border-radius:4px;font-size:12px;cursor:pointer;white-space:nowrap;vertical-align:middle;">'
            '&#x21BB; Recalculate</button>',
            uid=uid, pdf=pdf_url, excel=excel_url, csv=csv_url, recalc=recalc_url,
        )
    ledger_actions.short_description = "Actions"
    ledger_actions.allow_tags = True

    # ─── inlines ───────────────────────────────────────────────────────────

    inlines = [CustomerLedgerTransactionsInline, CustomerLedgerRequestInline]

    change_form_template = 'admin/customers/customerledger/change_form.html'

    # ─── custom URLs ───────────────────────────────────────────────────────

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'accept-customer-request/<int:request_id>/',
                self.admin_site.admin_view(self.accept_customer_request),
                name='accept_customer_request'
            ),
            path(
                'reject-customer-request/<int:request_id>/',
                self.admin_site.admin_view(self.reject_customer_request),
                name='reject_customer_request'
            ),
            path(
                'customer-ledger/<int:ledger_id>/recalculate/',
                self.admin_site.admin_view(self.recalculate_balance),
                name='recalculate_customer_balance',
            ),
            path(
                'customer-ledger/<int:ledger_id>/pdf/',
                self.admin_site.admin_view(self.download_ledger_pdf),
                name='download_customer_ledger',
            ),
            path(
                'customer-ledger/<int:ledger_id>/excel/',
                self.admin_site.admin_view(self.download_ledger_excel),
                name='download_customer_ledger_excel',
            ),
            path(
                'customer-ledger/<int:ledger_id>/csv/',
                self.admin_site.admin_view(self.download_ledger_csv),
                name='download_customer_ledger_csv',
            ),
        ]
        return custom_urls + urls

    # ─── existing request views ────────────────────────────────────────────

    def accept_customer_request(self, request, request_id):
        customer_request = CustomerRequestTransaction.objects.get(id=request_id)
        try:
            with transaction.atomic():
                customerLedgerEntry = CustomerLedgerTransaction.objects.create(
                    customer_ledger=customer_request.customer_ledger,
                    payment_type=customer_request.payment_type,
                    mode=customer_request.mode,
                    paid_on=customer_request.paid_on,
                    amount=customer_request.amount,
                    paid_to=customer_request.paid_to,
                    detail=customer_request.detail,
                )
                if customerLedgerEntry:
                    customer_request.status = CustomerRequestTransaction.STATUS_ACCEPTED
                    customer_request.save()

            self.message_user(request, "Customer request accepted successfully.", level='success')
        except Exception as e:
            self.message_user(request, f"Error: {str(e)}", level='error')

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    def reject_customer_request(self, request, request_id):
        customer_request = CustomerRequestTransaction.objects.get(id=request_id)
        customer_request.status = CustomerRequestTransaction.STATUS_DELETED
        customer_request.save()
        self.message_user(request, "Customer request rejected successfully.", level='success')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    def recalculate_balance(self, request, ledger_id):
        try:
            ledger = CustomerLedger.objects.get(pk=ledger_id)
            credited = CustomerLedgerTransaction.objects.filter(
                customer_ledger=ledger, payment_type='credited'
            ).aggregate(total=Sum('amount'))['total'] or 0
            debited = CustomerLedgerTransaction.objects.filter(
                customer_ledger=ledger, payment_type='debited'
            ).aggregate(total=Sum('amount'))['total'] or 0
            ledger.balance = ledger.amount - credited + debited
            ledger.save(update_fields=['balance'])
            self.message_user(
                request,
                f"Balance recalculate ho gaya: {format_indian_currency(ledger.balance)}",
                level='success'
            )
        except CustomerLedger.DoesNotExist:
            self.message_user(request, "Ledger nahi mila.", level='error')
        except Exception as e:
            self.message_user(request, f"Error: {str(e)}", level='error')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '../'))

    # ─── download helpers & views ──────────────────────────────────────────

    def _get_ledger_context(self, ledger_id):
        ledger = CustomerLedger.objects.select_related().get(pk=ledger_id)
        transactions = list(
            CustomerLedgerTransaction.objects.filter(customer_ledger=ledger)
            .select_related('paid_to')
            .order_by(F('paid_on').asc(nulls_last=True))
        )
        for txn in transactions:
            txn.formatted_amount = format_indian_currency(txn.amount)
        paid_amount = (ledger.amount - ledger.balance) if (ledger.amount and ledger.balance) else 0
        paid_words  = num2words(int(paid_amount), lang='en_IN').title() if paid_amount else ""
        return ledger, transactions, paid_amount, paid_words

    def download_ledger_pdf(self, request, ledger_id):
        ledger, transactions, paid_amount, paid_words = self._get_ledger_context(ledger_id)
        context = {
            'ledger':            ledger,
            'transactions':      transactions,
            'paid_amount':       paid_amount,
            'paid_words':        paid_words,
            'formatted_amount':  format_indian_currency(ledger.amount),
            'formatted_balance': format_indian_currency(ledger.balance),
            'formatted_paid':    format_indian_currency(paid_amount),
        }
        if weasyprint is None:
            return HttpResponse("PDF generation is unavailable (weasyprint not installed).", status=503)
        html_string = render_to_string('admin/customers/customer_ledger_pdf.html', context)
        pdf_file    = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="ledger_{ledger_id}.pdf"'
        return response

    def download_ledger_excel(self, request, ledger_id):
        if openpyxl is None:
            return HttpResponse("Excel generation is unavailable (openpyxl not installed).", status=503)

        ledger, transactions, paid_amount, _ = self._get_ledger_context(ledger_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Ledger"

        header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill('solid', fgColor='1A56DB')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        thin_border  = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin'),
        )
        title_font  = Font(name='Calibri', bold=True, size=14, color='1A56DB')
        label_font  = Font(name='Calibri', bold=True, size=10)
        value_font  = Font(name='Calibri', size=10)
        blue_fill   = PatternFill('solid', fgColor='EFF6FF')

        customer = ledger.customer_id
        customer_name = f"{customer.first_name} {customer.last_name}".strip() or str(customer)

        # Row 1: Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "SVED — Customer Ledger Statement"
        ws['A1'].font      = title_font
        ws['A1'].alignment = center_align

        # Rows 3+: Ledger info
        info = [
            ("Customer",      customer_name),
            ("Project",       str(ledger.project_id) if ledger.project_id else "—"),
            ("House",         str(ledger.project_house_id) if ledger.project_house_id else "—"),
            ("Total Amount",  format_indian_currency(ledger.amount)),
            ("Paid Amount",   format_indian_currency(paid_amount)),
            ("Balance",       format_indian_currency(ledger.balance)),
        ]
        for i, (lbl, val) in enumerate(info, start=3):
            ws.cell(row=i, column=1, value=lbl).font = label_font
            ws.cell(row=i, column=1).fill            = blue_fill
            ws.cell(row=i, column=2, value=val).font = value_font
            ws.cell(row=i, column=2).alignment       = left_align

        # Row 10: Transaction table header
        txn_header_row = 10
        txn_cols = ['#', 'Paid On', 'Payment Type', 'Paid To', 'Detail', 'Amount (₹)']
        for col_idx, col_name in enumerate(txn_cols, start=1):
            cell = ws.cell(row=txn_header_row, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        # Transaction rows
        for row_idx, txn in enumerate(transactions, start=1):
            row_num  = txn_header_row + row_idx
            row_fill = PatternFill('solid', fgColor='F5F8FF') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            paid_to_name = ""
            if txn.paid_to:
                paid_to_name = f"{txn.paid_to.first_name} {txn.paid_to.last_name}".strip() or str(txn.paid_to)
            row_data = [
                row_idx,
                txn.paid_on.strftime('%d-%m-%Y') if txn.paid_on else '—',
                txn.payment_type or '—',
                paid_to_name or '—',
                txn.detail or '—',
                float(txn.amount) if txn.amount else 0,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.border    = thin_border
                cell.alignment = center_align if col_idx in (1, 2, 3, 6) else left_align

        # Column widths
        col_widths = [6, 14, 16, 20, 36, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="customer_ledger_{ledger_id}.xlsx"'
        return response

    def download_ledger_csv(self, request, ledger_id):
        ledger, transactions, paid_amount, _ = self._get_ledger_context(ledger_id)

        customer = ledger.customer_id
        customer_name = f"{customer.first_name} {customer.last_name}".strip() or str(customer)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="customer_ledger_{ledger_id}.csv"'
        response.write('﻿')  # UTF-8 BOM

        writer = csv.writer(response)
        writer.writerow(['SVED - Customer Ledger Statement'])
        writer.writerow([])
        writer.writerow(['Customer',     customer_name])
        writer.writerow(['Project',      str(ledger.project_id) if ledger.project_id else '—'])
        writer.writerow(['House',        str(ledger.project_house_id) if ledger.project_house_id else '—'])
        writer.writerow(['Total Amount', format_indian_currency(ledger.amount)])
        writer.writerow(['Paid Amount',  format_indian_currency(paid_amount)])
        writer.writerow(['Balance',      format_indian_currency(ledger.balance)])
        writer.writerow([])

        writer.writerow(['#', 'Paid On', 'Payment Type', 'Paid To', 'Detail', 'Amount'])
        for i, txn in enumerate(transactions, start=1):
            paid_to_name = ""
            if txn.paid_to:
                paid_to_name = f"{txn.paid_to.first_name} {txn.paid_to.last_name}".strip() or str(txn.paid_to)
            writer.writerow([
                i,
                txn.paid_on.strftime('%d-%m-%Y') if txn.paid_on else '—',
                txn.payment_type or '—',
                paid_to_name or '—',
                txn.detail or '—',
                float(txn.amount) if txn.amount else 0,
            ])

        return response

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        if object_id:
            extra_context['download_pdf_url']   = reverse('admin:download_customer_ledger',       args=[object_id])
            extra_context['download_excel_url']  = reverse('admin:download_customer_ledger_excel', args=[object_id])
            extra_context['download_csv_url']    = reverse('admin:download_customer_ledger_csv',   args=[object_id])
        return super().changeform_view(request, object_id, form_url, extra_context)

    # ─── filters / search ──────────────────────────────────────────────────

    search_fields = ['amount', 'balance']
    list_filter   = [TableForiegnKeyListFilter("Projects", "project_id", "name", Projects)]

    class Media:
        js = ('admin/js/amountFormat1.js', 'admin/js/ledger_dropdown.js')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.none()


admin.site.register(CustomerLedger, CustomerLedgerAdmin)